import pandas as pd
import openpyxl
import os

print("🔄 Integrando productos respetando la estructura original...")

archivo_base = "BASE DE DATOS.xlsx"
archivo_nuevos = "Nuevos_Productos_Internet.xlsx"

if not os.path.exists(archivo_base) or not os.path.exists(archivo_nuevos):
    print("❌ Error: Falta alguno de los archivos.")
else:
    # Cargamos la base maestra conservando todas sus columnas originales
    wb = openpyxl.load_workbook(archivo_base)
    ws = wb.active

    # Leemos los datos actuales usando pandas para comparar códigos
    df_base = pd.read_excel(archivo_base, dtype=str)
    df_nuevos = pd.read_excel(archivo_nuevos, dtype=str)

    codigos_existentes = set(df_base['Código'].astype(str))
    
    # Filtramos solo los productos que verdaderamente no están en la base
    df_a_insertar = df_nuevos[~df_nuevos['Código'].astype(str).isin(codigos_existentes)].copy()

    if len(df_a_insertar) > 0:
        # Obtenemos las cabeceras exactas de tu base maestra
        headers = [cell.value for cell in ws[1]]
        
        # Añadimos fila por fila respetando las columnas de la base original
        for _, row in df_a_insertar.iterrows():
            nueva_fila = []
            for col in headers:
                # Si la columna existe en el nuevo, la agregamos; si no, va vacía para mantener el formato
                valor = row[col] if col in df_nuevos.columns else ""
                nueva_fila.append(str(valor) if valor is not None else "")
            
            ws.append(nueva_fila)

        # Blindamos la columna de códigos con formato de texto estricto (@)
        for cell in ws['A']:
            if cell.row > 1:
                cell.number_format = '@'
                cell.data_type = 's'

        wb.save(archivo_base)
        print(f"✅ ¡Integración limpia! Se agregaron {len(df_a_insertar)} productos manteniendo intacto el formato.")
    else:
        print("ℹ️ No hay productos nuevos que agregar. Tu base está impecable.")