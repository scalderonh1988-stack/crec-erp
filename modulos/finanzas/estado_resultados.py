import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

print("📊 Consolidando Estado de Resultados (P&L) del Negocio...")

# Archivos de nuestro ERP
archivo_ventas = "Ventas_Diarias.xlsx"
archivo_gastos = "Gastos_Negocio.xlsx"
archivo_cxp = "Cuentas_Por_Pagar.xlsx"
archivo_reporte = "Reporte_Estado_Resultados.xlsx"

# 1. Calculamos Ventas Totales
total_ventas_brutas = 0.0
if os.path.exists(archivo_ventas):
    df_v = pd.read_excel(archivo_ventas)
    if 'Total_Venta' in df_v.columns:
        total_ventas_brutas = df_v['Total_Venta'].sum()

# Desglose aproximado de IVA (19%) para obtener Venta Neta
IVA = 0.19
total_ventas_netas = total_ventas_brutas / (1 + IVA)
iva_debito = total_ventas_brutas - total_ventas_netas

# 2. Calculamos Gastos Operativos Totales
total_gastos = 0.0
if os.path.exists(archivo_gastos):
    df_g = pd.read_excel(archivo_gastos)
    if 'Monto' in df_g.columns:
        total_gastos = df_g['Monto'].sum()

# 3. Calculamos Cuentas por Pagar / Proveedores Totales
total_cxp = 0.0
if os.path.exists(archivo_cxp):
    df_c = pd.read_excel(archivo_cxp)
    if 'Monto_Total' in df_c.columns and 'Estado' in df_c.columns:
        pendientes = df_c[df_c['Estado'].str.upper() == 'PENDIENTE']
        total_cxp = pendientes['Monto_Total'].sum()

# Estimamos Costo de Ventas (COGS) proporcional o basado en margen estándar (ej. 70% de ventas netas si no hay registro detallado, o 0)
# Para este reporte gerencial, estructuraremos la P&L limpia:
utilidad_bruta = total_ventas_netas  # Como base de ingresos netos operacionales
utilidad_operativa = utilidad_bruta - total_gastos

# 4. Generamos el reporte profesional en Excel con openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Estado_Resultados"

# Aseguramos que se muestren las cuadrículas en Excel
ws.views.sheetView[0].showGridLines = True

# Estilos profesionales (Paleta Slate / Azul Oscuro corporativo)
font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
font_section = Font(name="Calibri", size=12, bold=True, color="1F497D")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
font_normal = Font(name="Calibri", size=11, color="000000")

fill_title = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
fill_total = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
double_bottom_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='double', color='000000')
)

# Título Principal
ws.merge_cells("A1:C1")
cell_title = ws["A1"]
cell_title.value = "ESTADO DE RESULTADOS (P&L) - RESUMEN FINANCIERO"
cell_title.font = font_title
cell_title.fill = fill_title
cell_title.alignment = align_center
ws.row_dimensions[1].height = 35

# Subtítulo con fecha de emisión
ws.merge_cells("A2:C2")
cell_sub = ws["A2"]
cell_sub.value = f"Fecha de emisión: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
cell_sub.font = Font(name="Calibri", size=10, italic=True, color="595959")
cell_sub.alignment = align_center
ws.row_dimensions[2].height = 20

# Encabezados de Tabla
ws.append([]) # Fila 3 en blanco
headers = ["Concepto Financiero", "Detalle / Referencia", "Monto ($)"]
ws.append(headers)
ws.row_dimensions[4].height = 25

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Datos del Estado de Resultados
filas_datos = [
    ("(+) Ventas Brutas Totales (Caja)", "Total recaudado en terminal de ventas", total_ventas_brutas),
    ("(-) IVA Débito Fiscal (Estimado 19%)", "Impuesto a las ventas retenido", iva_debito),
    ("(=) Ventas Netas del Período", "Ingresos netos operacionales", total_ventas_netas),
    ("(-) Gastos Operativos Totales", "Egresos, servicios y gastos generales", total_gastos),
    ("(=) Utilidad Operativa Estimada", "Ganancia antes de compromisos de proveedores", utilidad_operativa),
    ("(ℹ) Cuentas por Pagar (Proveedores)", "Deuda total pendiente con proveedores", total_cxp)
]

for row_data in filas_datos:
    ws.append(list(row_data))
    current_row = ws.max_row
    ws.row_dimensions[current_row].height = 20
    
    # Aplicar estilos según el tipo de fila
    is_total = "=" in row_data[0]
    is_info = "ℹ" in row_data[0]
    
    for col_idx in range(1, 4):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.border = thin_border
        if col_idx == 3:
            cell.number_format = '$#,##0.00'
            cell.alignment = align_right
        else:
            cell.alignment = align_left
            
        if is_total:
            cell.font = font_bold
            cell.fill = fill_total
        elif is_info:
            cell.font = Font(name="Calibri", size=11, italic=True, color="333333")
        else:
            cell.font = font_normal

# Ajuste automático de ancho de columnas
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max(max_len + 5, 25)

wb.save(archivo_reporte)

print(f"\n=========================================================")
print(f"✅ ¡Estado de Resultados (P&L) generado con éxito!")
print(f"📁 Archivo guardado profesionalmente como: '{archivo_reporte}'.")
print(f"---------------------------------------------------------")
print(f"💵 Ventas Netas: ${total_ventas_netas:,.2f}")
print(f"📉 Gastos Operativos: ${total_gastos:,.2f}")
print(f"💰 Utilidad Operativa: ${utilidad_operativa:,.2f}")
print(f"📋 Deuda Pendiente Proveedores: ${total_cxp:,.2f}")
print("=========================================================")