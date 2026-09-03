import pandas as pd
import openpyxl
import os
from datetime import datetime

print("🔧 Verificando y Actualizando Base de Datos de Mermas...")

archivo_base = "BASE DE DATOS.xlsx"
archivo_mermas = "Registro_Mermas.xlsx"

# Si no existe el registro de mermas, creamos la estructura base
if not os.path.exists(archivo_mermas):
    df_inicial = pd.DataFrame(columns=[
        'Fecha_Hora',
        'Codigo',
        'Descripcion',
        'Cantidad_Perdida',
        'Costo_Unitario',
        'Perdida_Total',
        'Motivo'
    ])
    df_inicial.to_excel(archivo_mermas, index=False)

if not os.path.exists(archivo_base):
    print(f"❌ Error crítico: No se encuentra el archivo maestro '{archivo_base}'.")
else:
    # 1. Cargamos con pandas asegurando que el código sea texto
    df_base = pd.read_excel(archivo_base, dtype={'Código': str})
    
    # Verificamos si existe la columna de Stock; si no, la creamos automáticamente
    col_stock = next((col for col in df_base.columns if 'stock' in str(col).lower() or 'cantidad' in str(col).lower() or 'existencia' in str(col).lower()), None)
    
    if not col_stock:
        print("⚠️ No se encontró la columna de Stock en tu base de datos.")
        print("🛠️ Creando automáticamente la columna 'Stock' con un valor inicial de 50 unidades...")
        df_base['Stock'] = 50.0
        df_base.to_excel(archivo_base, index=False)
        col_stock = 'Stock'

    # Cargamos con openpyxl para manipular celdas exactas
    wb = openpyxl.load_workbook(archivo_base)
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value is not None]
    
    # Aseguramos que la columna 'Stock' esté reflejada en el openpyxl
    if col_stock not in headers:
        ws.cell(row=1, column=len(headers) + 1, value='Stock')
        wb.save(archivo_base)
        headers = [cell.value for cell in ws[1] if cell.value is not None]

    idx_stock_ws = headers.index(col_stock) + 1
    col_costo = next((col for col in df_base.columns if 'costo' in str(col).lower() or 'compra' in str(col).lower()), None)

    print("\n---------------------------------------------------------")
    print("🔴 REGISTRO DE MERMAS Y PRODUCTOS DAÑADOS")
    print("---------------------------------------------------------")
    print("Instrucciones: Ingresa el código del producto mermado (o 'salir').")

    mermas_sesion = []

    while True:
        codigo_ingresado = input("\nCódigo del producto a dar de baja (o 'salir'): ").strip()

        if codigo_ingresado.lower() == 'salir':
            break

        if not codigo_ingresado:
            continue

        # Buscamos el producto por código EAN o interno
        match = df_base[df_base['Código'].astype(str).str.strip() == codigo_ingresado]

        if match.empty:
            print(f"❌ Producto con código '{codigo_ingresado}' no encontrado.")
        else:
            row_index_df = match.index[0]
            descripcion = match['Descripción'].values[0] if 'Descripción' in match.columns else "Sin descripción"
            
            # Conversión segura del stock
            val_stock_raw = match[col_stock].values[0]
            try:
                stock_actual = float(val_stock_raw) if pd.notna(val_stock_raw) else 0.0
            except (ValueError, TypeError):
                stock_actual = 0.0

            # Conversión segura del costo
            costo_unitario = 0.0
            if col_costo and pd.notna(match[col_costo].values[0]):
                try:
                    costo_unitario = float(match[col_costo].values[0])
                except (ValueError, TypeError):
                    costo_unitario = 0.0

            print(f"📦 Producto: {descripcion}")
            print(f"📊 Stock Actual: {stock_actual} | Costo Unitario: ${costo_unitario:,.2f}")

            try:
                cantidad_merma = float(input("Cantidad que se pierde / desecha: ") or "0")
            except ValueError:
                cantidad_merma = 0.0

            if cantidad_merma <= 0:
                print("⚠️ Cantidad inválida.")
                continue

            if cantidad_merma > stock_actual:
                print("⚠️ Advertencia: Estás mermando más de lo que figura en el stock actual.")

            motivo = input("Motivo [Vencido / Roto / Robado / Consumo Interno / Otro]: ").strip() or "Vencido"

            # Actualizamos stock
            nuevo_stock = stock_actual - cantidad_merma
            df_base.loc[row_index_df, col_stock] = nuevo_stock

            row_ws_idx = row_index_df + 2
            ws.cell(row=row_ws_idx, column=idx_stock_ws, value=nuevo_stock)

            perdida_total_linea = cantidad_merma * costo_unitario

            mermas_sesion.append({
                'Fecha_Hora': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Codigo': codigo_ingresado,
                'Descripcion': descripcion,
                'Cantidad_Perdida': cantidad_merma,
                'Costo_Unitario': costo_unitario,
                'Perdida_Total': perdida_total_linea,
                'Motivo': motivo
            })

            print(f"✅ Merma registrada. Pérdida: ${perdida_total_linea:,.2f} | Nuevo Stock: {nuevo_stock}")

    # Guardamos cambios finales
    if mermas_sesion:
        # Asegurar formato de texto en la columna de códigos al guardar
        for cell in ws['A']:
            if cell.row > 1:
                cell.number_format = '@'

        wb.save(archivo_base)

        df_nuevas_mermas = pd.DataFrame(mermas_sesion)
        df_mermas_antiguas = pd.read_excel(archivo_mermas)
        df_mermas_final = pd.concat([df_mermas_antiguas, df_nuevas_mermas], ignore_index=True)
        df_mermas_final.to_excel(archivo_mermas, index=False)

        total_perdida_sesion = df_nuevas_mermas['Perdida_Total'].sum()
        print("\n---------------------------------------------------------")
        print("🏁 REGISTRO DE MERMAS FINALIZADO")
        print(f"📁 Stock actualizado y mermas guardadas en '{archivo_mermas}'.")
        print(f"📉 Pérdida total acumulada en esta sesión: ${total_perdida_sesion:,.2f}")
        print("---------------------------------------------------------")
    else:
        print("\nℹ️ No se registraron mermas en esta sesión.")